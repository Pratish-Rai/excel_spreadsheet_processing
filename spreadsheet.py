import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        enter_value = cell.value * 10
        enter_value_cell = sheet.cell(row, 2)
        enter_value_cell.value = enter_value
        values = Reference(sheet,
                           min_row=1,
                           max_row=sheet.max_row,
                           min_col=2,
                           max_col=2)
        chart = BarChart3D()
        chart.add_data(values)
        sheet.add_chart(chart, "a15")
        wb.save(filename)


process_workbook("monkey.xlsx")